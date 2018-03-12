import sys
from PyQt5.QtWidgets import QWidget, QApplication, QHBoxLayout, QFrame, QVBoxLayout, QLabel, QLineEdit, QPushButton, QTreeView, QCheckBox, QGridLayout, QMessageBox
from excelio import ExcelQtConverter
from PyQt5.QtGui import QStandardItem, QStandardItemModel
from PyQt5.QtCore import Qt
import openpyxl
from datetime import datetime
#import pdb; pdb.set_trace()


class MainWindow(QWidget):
    '''
    メインウィンドウとして
    実はこれしかウィンドウ無い
    '''

    def __init__(self):
        '''
        メインウィンドウの設定
        '''
        super().__init__()
        self.init_UI()
        self.make_cart_model()

        #カートに追加するための操作。予めExcelファイルを開いておく方が、add_to_cartで毎回開くより早いことが判明。
        self.cart_row = 0
        self.book = openpyxl.load_workbook('Python リサイクル市 会計用 Er.xlsx', data_only = True)
        self.readsheet = self.book['raw']
        self.writesheet = self.book['会計録']
        self.customersheet = self.book['customer_no']

        self.total_price = 0
        self.set_customer_number()

    def init_UI(self):
        self.setWindowTitle('読込テスト2')

        
        
        outer_layout = QHBoxLayout()

        #ここから左カラムの作成
        left_container = QFrame()
        left_container.setFrameStyle(1)
        left_container.setFrameShadow(QFrame.Sunken)
        left = QVBoxLayout(left_container)

        #ラベル作成
        lbl11 = QLabel("商品番号",self)
        lbl12 = QLabel("値引き後の値段\r\n（初期価格は空白）", self)

        #入力バー作成
        self.txtbox11 = QLineEdit(self)
        self.txtbox11.returnPressed.connect(self.add_to_cart)
        self.txtbox12 = QLineEdit(self)
        self.txtbox12.returnPressed.connect(self.add_to_cart)

        #ボタン作成
        add_to_cart_button = QPushButton("カートに追加",self)
        add_to_cart_button.clicked.connect(self.add_to_cart)

        #leftのwidget配置
        left.addWidget(lbl11)
        left.addWidget(self.txtbox11)
        left.addStretch(1)
        left.addWidget(lbl12)
        left.addWidget(self.txtbox12)
        left.addStretch(2)
        left.addWidget(add_to_cart_button)
        left.addStretch(10)


        #ここから真ん中のカラムの作成
        middle_container = QFrame()
        middle_container.setFrameStyle(1)
        middle_container.setFrameShadow(QFrame.Sunken) 
        middle = QVBoxLayout(middle_container)
        
        #middleのカート
        lbl21 = QLabel("カート", self)

        #ツリー
        self.cart_view = QTreeView(self)

        reset_cart_button = QPushButton("全てクリア", self)
        reset_cart_button.clicked.connect(self.reset_cart)

        #middleのチェックボックス
        lbl22 = QLabel("配送", self)
        self.check21 = QCheckBox("配送(\\500)", self)
        self.check21.stateChanged.connect(self.select_delivery)
        
        #middleのwidget配置
        middle.addWidget(lbl21)
        middle.addWidget(self.cart_view)
        #middle.addStretch(1)
        middle.addWidget(reset_cart_button)
        middle.addWidget(lbl22)
        middle.addWidget(self.check21)
         
       
        #ここから右カラムの作成
        right_container = QFrame()
        right_container.setFrameStyle(1)
        right_container.setFrameShadow(QFrame.Sunken)
        right = QGridLayout(right_container)

        #ラベル作成
        lbl1 = QLabel("値段",self)
        lbl2 = QLabel("お預り", self)
        lbl3 = QLabel("おつり", self)
        self.lbl4 = QLabel("",self)

        # 入力バー作成
        self.txtbox1 = QLabel(self)
        self.txtbox1.setText("0")
        self.txtbox2 = QLineEdit(self)

        # OKボタン作成
        OKbutton = QPushButton("OK", self)
        OKbutton.clicked.connect(self.calc_on_click)

        # 記録ボタン作成
        LOGbutton = QPushButton("記帳", self)
        LOGbutton.clicked.connect(self.write_to_Excel)

        #rightのwidget配置
        right.setRowStretch(0, 1)
        right.setRowStretch(4, 1)
        right.addWidget(lbl1, 1, 0)
        right.addWidget(lbl2, 2, 0)
        right.addWidget(lbl3, 3, 0)
        right.addWidget(self.txtbox1,1,1)
        right.addWidget(self.txtbox2,2,1)
        right.addWidget(self.lbl4, 3, 1)        
        right.addWidget(OKbutton,3,2)
        right.addWidget(LOGbutton,5,2)


        #大枠の配置
        outer_layout.addWidget(left_container)
        outer_layout.addWidget(middle_container)
        outer_layout.addWidget(right_container)

        self.setLayout(outer_layout)


        self.show()



##    def make_full_item_list(self):
##        #エクセルファイルのrawをQStandardItemModelに変換するよ！       
##        cvtr = ExcelQtConverter('Python リサイクル市 会計用.xlsx')
##        self.inventory = cvtr.to_model('raw')
##        self.item_list.setModel(self.inventory)
        
        

    def make_cart_model(self):
        self.cart_model = QStandardItemModel()
        self.cart_view.setModel(self.cart_model)
        self.cart_model.setHorizontalHeaderLabels(["商品番号", "商品名", "価格"])
        
    def calc_on_click(self):
        try:
            p = int(self.txtbox1.text())
            m = int(self.txtbox2.text())
            change = Calc(p, m)
            self.lbl4.setText(str(change))
            print(p,m,change)
        except:
            pass

    def select_delivery(self,state):
        if Qt.Checked == state:
            self.total_price += 500
        elif Qt.Unchecked == state:
            self.total_price -= 500

        self.txtbox1.setText(str(self.total_price))
            
        

##    def add_to_cart(self):
##        self.item_num = self.txtbox11.text()
##        qt_item1 = QStandardItem()
##        qt_item2 = QStandardItem()
##        qt_item3 = QStandardItem()
##        qt_item1.setText(self.item_num)
##        qt_item2.setText(self.product_name)
##        qt_item3.setText(self.product_price)
##        self.cart_model.setItem(0, 0, qt_item1)
##        self.cart_model.setItem(0, 1, qt_item2)
##        self.cart_model.setItem(0, 2, qt_item3)
        
        
    def add_to_cart(self):

        try:
            number = int(self.txtbox11.text())

        except:
            if not self.txtbox11.text() == "":
                disabled_dialog = QMessageBox.information(self, 'エラー 無効な入力', '半角数字を入力してください', QMessageBox.Ok)

        else:
            #読込2では、__init__した時点でExcelファイルを開いて少しでもレジ打ち動作を早めようとしている。
            #そのため、位置だけでなく変数にself.を加えるなど若干の変数名変更がある。
            
            
            #入力された商品番号をExcelファイルから直接検索する
            

            #ヘッダーを除いて順番に読み込む
            for row in range(2, self.readsheet.max_row+1):
            
            #本当は直接、if self.product_id == self.item_numにして、商品番号と一致した時点で読み込みをbreakして
            #その時点におけるproduct_nameとproduct_priceをQStandardItemに変換して表示できればいいんだけど、
            #どうも一度number = int(self.item_num)を挟まないと上手くbreakが行われず、
            #Excelファイルの最後の商品（ホーロー容器01）が表示されてしまう。
            #Excelファイルの最後の行にエラー番号を付加し、
            #実在しない商品番号(桁は不問)を打ち込むとコンソールにエラーを吐いてsetItemされない。clearはされる。
            #ただ、数字以外を打ち込まれると固まる。
            #というかなぜか価格はバグで表示されない（価格をコメントアウトしないと動作停止する）
            #あと、順番に読み込んでいるためか動作が遅い。最初にExcelファイルを開いておくことで多少は改善された。

                self.product_id = self.readsheet['A' +str(row)].value
                self.product_name = self.readsheet['B' +str(row)].value
                self.product_price = self.readsheet['C' +str(row)].value

                #print(self.product_id,self.product_name,self.product_price)#確認用としてコンソールにprintしてもらう(時間かかる)。価格もここまでは大丈夫
                
                if self.product_id == 99999:
                    
                    error_dialog = QMessageBox.information(self, 'エラー 該当なし', '該当商品が見つかりません', QMessageBox.Ok)

                    
    
                    
                elif self.product_id == number:
                    
                    qt_item1 = QStandardItem()
                    qt_item2 = QStandardItem()
                    qt_item3 = QStandardItem()

                    qt_item1.setText(self.txtbox11.text())
                    qt_item2.setText(self.product_name)

                    if self.txtbox12.text() == "":
                        qt_item3.setText(str(self.product_price))  # ←ここでエラー出るのは、strで指定していないため。
                        # qt_item3.setText(self.product_name)
                    else:
                        self.product_price = int(self.txtbox12.text())
                        qt_item3.setText(str(self.product_price))

                    self.cart_model.setItem(self.cart_row, 0, qt_item1)
                    self.cart_model.setItem(self.cart_row, 1, qt_item2)
                    self.cart_model.setItem(self.cart_row, 2, qt_item3)

                    self.cart_row += 1


                    self.total_price += self.product_price
                    self.txtbox1.setText(str(self.total_price))
                    
                    
                    break
                    
        self.txtbox11.clear()
        self.txtbox12.clear()

                #self.cart_model_setItemで、座標を(0,0,qt_item1)と指定すると次々その部分に上書きされてしまうため、
                #__init__で適当な変数を用意して座標とし、アイテムを一つsetItemする毎にその変数を一つずつ大きくしている。
                #ついでに入力欄をclear(＝消去)している          


    

    def reset_cart(self):
        self.cart_model.clear()
        self.cart_row = 0
        self.cart_model.setHorizontalHeaderLabels(["商品番号", "商品名", "価格"])
        self.check21.setChecked(False)
        self.total_price = 0
        self.txtbox1.setText(str(self.total_price))
        self.txtbox11.clear()
        self.txtbox12.clear()
        self.txtbox2.clear()
        self.lbl4.setText("")
        
    def write_to_Excel(self):
        newrow = self.writesheet.max_row + 1
        #max_row は一度入力されdelで値が削除されたセルも使用済みと認識するため
        #手動で訂正する場合は列を削除する必要がある。
        self.writesheet['A' + str(newrow)] = int(self.customer_number)
        self.writesheet['B' + str(newrow)] = str(datetime.today())
        newrow_for_items = newrow
        for i in range(0, int(self.cart_model.rowCount())):
            self.writesheet['C' + str(newrow_for_items)] = int(self.cart_model.item(i, 0).text())
            self.writesheet['D' + str(newrow_for_items)] = self.cart_model.item(i, 1).text()
            self.writesheet['E' + str(newrow_for_items)] = int(self.cart_model.item(i, 2).text())
            newrow_for_items += 1

        if self.check21.isChecked():
            self.writesheet['C' + str(newrow_for_items)] = 'NONE'
            self.writesheet['D' + str(newrow_for_items)] = '配送料'
            self.writesheet['E' + str(newrow_for_items)] = 500
            newrow_for_items += 1
        else:
            pass
            
        self.customersheet['A' + str(self.customer_number + 1)] = self.customer_number
        self.customer_number += 1
        self.book.save('Python リサイクル市 会計用 Er.xlsx')
        self.reset_cart()
        #書き込む場合はExcelファイルを閉じておくように。


    def set_customer_number(self):
        #会計録シートのA列の最後に使用されているセルの番号を読取るコードを書きたかったが、
        #会計録シートで列の最後に使用されているセルを取得するlen(ws[row])では
        #書込時使用された他の列の行の最大値になってしまうため新しくシートを増やした。
        #手動で修正する場合は列の削除をしなければならない。
        #改善の余地あり...
        row_len = len(self.customersheet['A'])
        if row_len == 1:
            self.customer_number = 1
        else:
            self.customer_number = self.customersheet['A' +str(row_len)].value + 1

                

        
    '''
    def pass_on_click(self):
        item_num = self.txtbox11.text()
        qt_item = QStandardItem()
        qt_item.setText(item_num)
        self.inventory.setItem(0, 0, qt_item)
    '''

        
def Calc(price, money):
    return money - price



##class Do_not_touch_me(QWidget):
##    #触らないで！と声に出そう
##
##    def __init__(self):
##        super().__init__()
##        self.setWindowTitle('触らないで！')
##        self.setGeometry(300, 200, 100, 150)
##        label = QLabel("触らないで！", self)
##
##        self.show()
        

app = QApplication(sys.argv)
main_window = MainWindow()
sys.exit(app.exec_())

